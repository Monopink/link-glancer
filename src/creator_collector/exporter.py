from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

EXPORT_HEADERS = [
    "creator_oecuid",
    "handle",
    "nickname",
    "selection_region",
    "units_sold",
    "follower_cnt",
    "ec_video_avg_view_cnt",
    "video_engagement",
    "ec_video_engagement",
    "video_avg_view_cnt",
    "category",
    "top_follower_gender",
    "url",
]

PROFILE_URL_PREFIX = "https://www.tiktok.com/@"


def export_collection_to_xlsx(rows: list[dict[str, object]], output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    rows = dedupe_creator_rows(rows)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Creators"

    for column, header in enumerate(EXPORT_HEADERS, start=1):
        sheet.cell(row=1, column=column, value=header)

    for row_index, source_row in enumerate(rows, start=2):
        flattened = flatten_creator_row(source_row)
        for column, header in enumerate(EXPORT_HEADERS, start=1):
            cell = sheet.cell(row=row_index, column=column, value=flattened.get(header, ""))
            if header == "creator_oecuid":
                cell.number_format = "@"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    export_path = _unique_path(output_dir / f"creator_{timestamp}.xlsx")
    workbook.save(export_path)
    workbook.close()
    return export_path


def export_collection_to_path(rows: list[dict[str, object]], export_path: Path) -> Path:
    export_path.parent.mkdir(parents=True, exist_ok=True)
    rows = dedupe_creator_rows(rows)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Creators"

    for column, header in enumerate(EXPORT_HEADERS, start=1):
        sheet.cell(row=1, column=column, value=header)

    for row_index, source_row in enumerate(rows, start=2):
        flattened = flatten_creator_row(source_row)
        for column, header in enumerate(EXPORT_HEADERS, start=1):
            cell = sheet.cell(row=row_index, column=column, value=flattened.get(header, ""))
            if header == "creator_oecuid":
                cell.number_format = "@"

    workbook.save(export_path)
    workbook.close()
    return export_path


def save_backup_json(rows: list[dict[str, object]], backup_path: Path) -> Path:
    backup_path.parent.mkdir(parents=True, exist_ok=True)
    backup_path.write_text(
        json.dumps(rows, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return backup_path


def dedupe_creator_rows(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    deduped: list[dict[str, object]] = []
    seen: set[str] = set()
    for row in rows:
        row_key = _creator_identity(row)
        if row_key in seen:
            continue
        seen.add(row_key)
        deduped.append(row)
    return deduped


def flatten_creator_row(item: dict[str, object]) -> dict[str, object]:
    result: dict[str, object] = {}

    creator_oecuid = _value_of(item.get("creator_oecuid"))
    handle = _value_of(item.get("handle"))

    result["creator_oecuid"] = creator_oecuid
    result["handle"] = handle
    result["nickname"] = _value_of(item.get("nickname"))
    result["selection_region"] = _value_of(item.get("selection_region"))
    result["units_sold"] = _value_of(item.get("units_sold"))
    result["follower_cnt"] = _value_of(item.get("follower_cnt"))
    result["ec_video_avg_view_cnt"] = _value_of(item.get("ec_video_avg_view_cnt"))
    result["video_engagement"] = _value_of(item.get("video_engagement"))
    result["ec_video_engagement"] = _value_of(item.get("ec_video_engagement"))
    result["video_avg_view_cnt"] = _value_of(item.get("video_avg_view_cnt"))
    result["category"] = _category_value(item.get("category"))
    result["top_follower_gender"] = _top_follower_gender_value(item.get("top_follower_gender"))
    result["url"] = f"{PROFILE_URL_PREFIX}{handle}" if handle else ""
    return result


def _value_of(raw: object) -> object:
    if not isinstance(raw, dict):
        return ""
    value = raw.get("value")
    if isinstance(value, dict):
        return value.get("minimal", "")
    if isinstance(value, list):
        return ", ".join(str(part) for part in value)
    return value if value is not None else ""


def _category_value(raw: object) -> str:
    if not isinstance(raw, dict):
        return ""
    value = raw.get("value")
    if not isinstance(value, list):
        return ""
    return ", ".join(
        str(item.get("name", "")) for item in value if isinstance(item, dict) and item.get("name")
    )


def _top_follower_gender_value(raw: object) -> str:
    if not isinstance(raw, dict):
        return ""
    value = raw.get("value")
    if not isinstance(value, list):
        return ""
    return ", ".join(
        f"{item.get('key', '')}: {item.get('value', '')}"
        for item in value
        if isinstance(item, dict)
    )


def _unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    for index in range(1, 10000):
        candidate = path.with_name(f"{path.stem}_{index:02d}{path.suffix}")
        if not candidate.exists():
            return candidate
    raise FileExistsError(f"Unable to create unique path for {path}")


def _creator_identity(row: dict[str, object]) -> str:
    creator_oecuid = str(_value_of(row.get("creator_oecuid")) or "").strip()
    handle = str(_value_of(row.get("handle")) or "").strip()
    if creator_oecuid:
        return creator_oecuid
    return f"{creator_oecuid}|{handle}"
