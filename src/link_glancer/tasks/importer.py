from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from link_glancer.tasks.models import TaskSnapshot


def read_workbook_sheet_names(workbook_path: Path) -> list[str]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def read_workbook_headers(workbook_path: Path, *, sheet_name: str, header_row: int) -> list[str]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet {sheet_name!r} not found in workbook")
        sheet = workbook[sheet_name]
        return _read_header_values(sheet, header_row)
    finally:
        workbook.close()


def import_task_workbook(
    workbook_path: Path, task_snapshot: TaskSnapshot
) -> list[tuple[int, dict[str, object]]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if task_snapshot.sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet {task_snapshot.sheet_name!r} not found in workbook")

        sheet = workbook[task_snapshot.sheet_name]
        header_map = _build_header_map(sheet, task_snapshot.header_row)
        rows: list[tuple[int, dict[str, object]]] = []
        if not header_map:
            return rows

        for row_number, values in _iter_data_rows(sheet, task_snapshot.header_row):
            row_values = {
                header: (values[column_index] if column_index < len(values) else None)
                for header, column_index in header_map.items()
            }
            if not any(value not in (None, "") for value in row_values.values()):
                continue
            normalized = {
                header: ("" if value is None else value)
                for header, value in row_values.items()
                if header.strip()
            }
            rows.append((row_number, normalized))

        return rows
    finally:
        workbook.close()


def workbook_headers_exist(
    workbook_path: Path, *, sheet_name: str, header_row: int, headers: list[str]
) -> tuple[list[str], list[str]]:
    existing_headers = read_workbook_headers(
        workbook_path, sheet_name=sheet_name, header_row=header_row
    )
    existing_normalized = {_normalize_header(header): header for header in existing_headers}
    found = [header for header in headers if _normalize_header(header) in existing_normalized]
    missing = [header for header in headers if _normalize_header(header) not in existing_normalized]
    return found, missing


def _build_header_map(sheet, header_row: int) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for column_index, header in enumerate(_read_header_values(sheet, header_row)):
        normalized = _normalize_header(header)
        if normalized and normalized not in header_map:
            header_map[header] = column_index
    return header_map


def _read_header_values(sheet, header_row: int) -> list[str]:
    for row in sheet.iter_rows(
        min_row=header_row,
        max_row=header_row,
        values_only=True,
    ):
        return [text for text in (str(cell or "").strip() for cell in row) if text]
    return []


def _iter_data_rows(sheet, header_row: int):
    yield from enumerate(
        sheet.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    )


def _normalize_header(value: str) -> str:
    return value.strip().casefold()
