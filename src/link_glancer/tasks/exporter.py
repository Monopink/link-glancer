from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from link_glancer.tasks.database import list_all_items, list_reviews, load_task_detail
from link_glancer.tasks.models import ReviewField, ReviewRecord, TaskItem


def export_task_results(
    database_path: Path, task_id: int, destination_dir: Path | None = None
) -> Path:
    task = load_task_detail(database_path, task_id)
    items = list_all_items(database_path, task_id)
    reviews_by_item_id = list_reviews(database_path, task_id)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Results"
    export_fields = _resolved_export_fields(
        task.task_snapshot.review_fields,
        task.task_snapshot.enabled_review_field_ids,
        items,
        task.task_snapshot.export_fields,
    )

    for column_index, export_field in enumerate(export_fields, start=1):
        sheet.cell(row=1, column=column_index, value=export_field)

    for row_index, item in enumerate(items, start=2):
        review = reviews_by_item_id.get(item.task_item_id)
        for column_index, export_field in enumerate(export_fields, start=1):
            value = _resolve_export_value(export_field, item, review)
            if isinstance(value, list):
                value = ", ".join(str(option) for option in value)
            sheet.cell(row=row_index, column=column_index, value=value)

    export_dir = destination_dir or task.source_file_path.parent
    export_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = _safe_filename(task.name)
    export_path = _unique_export_path(export_dir / f"{safe_name}_{timestamp}.xlsx")
    workbook.save(export_path)
    workbook.close()
    return export_path


def export_task_results_to_path(database_path: Path, task_id: int, export_path: Path) -> Path:
    task = load_task_detail(database_path, task_id)
    items = list_all_items(database_path, task_id)
    reviews_by_item_id = list_reviews(database_path, task_id)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Results"
    export_fields = _resolved_export_fields(
        task.task_snapshot.review_fields,
        task.task_snapshot.enabled_review_field_ids,
        items,
        task.task_snapshot.export_fields,
    )

    for column_index, export_field in enumerate(export_fields, start=1):
        sheet.cell(row=1, column=column_index, value=export_field)

    for row_index, item in enumerate(items, start=2):
        review = reviews_by_item_id.get(item.task_item_id)
        for column_index, export_field in enumerate(export_fields, start=1):
            value = _resolve_export_value(export_field, item, review)
            if isinstance(value, list):
                value = ", ".join(str(option) for option in value)
            sheet.cell(row=row_index, column=column_index, value=value)

    export_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(export_path)
    workbook.close()
    return export_path


def _resolve_export_value(field_name: str, item: TaskItem, review: ReviewRecord | None) -> object:
    if review and field_name in review.review_data:
        return review.review_data[field_name]
    if field_name in item.task_data:
        return item.task_data[field_name]
    return ""


def _resolved_export_fields(
    review_fields: list[ReviewField],
    enabled_review_field_ids: list[str],
    items: list[TaskItem],
    export_fields: list[str],
) -> list[str]:
    available_columns = {
        key.casefold()
        for item in items
        for key in item.task_data
        if isinstance(key, str) and key.strip()
    }
    enabled_ids = _enabled_review_field_ids(review_fields, enabled_review_field_ids)
    enabled_set = {field_id.casefold() for field_id in enabled_ids}
    resolved: list[str] = []
    seen: set[str] = set()

    for field_name in export_fields:
        normalized = field_name.casefold()
        if normalized in seen:
            continue
        if normalized in available_columns or normalized in enabled_set:
            resolved.append(field_name)
            seen.add(normalized)

    for field_id in enabled_ids:
        normalized = field_id.casefold()
        if normalized in seen:
            continue
        resolved.append(field_id)
        seen.add(normalized)
    return resolved


def _enabled_review_field_ids(
    review_fields: list[ReviewField],
    enabled_review_field_ids: list[str],
) -> list[str]:
    known_ids = [field.field_id for field in review_fields]
    if not enabled_review_field_ids:
        return known_ids
    enabled_set = {field_id for field_id in enabled_review_field_ids}
    return [field_id for field_id in known_ids if field_id in enabled_set]


def _safe_filename(value: str) -> str:
    sanitized = "".join(char if char not in '<>:"/\\|?*' else "_" for char in value).strip()
    return sanitized or "task_export"


def _unique_export_path(export_path: Path) -> Path:
    if not export_path.exists():
        return export_path

    stem = export_path.stem
    suffix = export_path.suffix
    for counter in range(1, 10000):
        candidate = export_path.with_name(f"{stem}_{counter:02d}{suffix}")
        if not candidate.exists():
            return candidate
    raise FileExistsError(f"Unable to create a unique export path for {export_path}")
